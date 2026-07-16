from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, selectinload
from typing import List
import uuid
from datetime import datetime
import os
import io

from app.database import get_db
from app.models.inventory import Supplier, StockReceipt, StockReceiptItem
from app.models.product import ProductSKU, Product
from app.schemas.inventory import (
    SupplierResponse, SupplierCreate, SupplierUpdate,
    StockReceiptResponse, StockReceiptCreate,
    InventorySKUResponse, SKUHistoryResponse
)
from app.routers.auth import get_current_admin
from app.models.user import User

router = APIRouter(
    prefix="/admin/inventory",
    tags=["Admin Inventory"],
    dependencies=[Depends(get_current_admin)]
)

# =======================
# SUPPLIERS
# =======================

@router.get("/suppliers", response_model=List[SupplierResponse])
def get_all_suppliers(db: Session = Depends(get_db)):
    return db.query(Supplier).order_by(Supplier.created_at.desc()).all()

@router.post("/suppliers", response_model=SupplierResponse, status_code=status.HTTP_201_CREATED)
def create_supplier(supplier_in: SupplierCreate, db: Session = Depends(get_db)):
    new_sup = Supplier(
        id=str(uuid.uuid4()),
        name=supplier_in.name,
        contact_name=supplier_in.contact_name,
        phone=supplier_in.phone,
        email=supplier_in.email,
        address=supplier_in.address,
        is_active=supplier_in.is_active
    )
    db.add(new_sup)
    db.commit()
    db.refresh(new_sup)
    return new_sup

@router.put("/suppliers/{sup_id}", response_model=SupplierResponse)
def update_supplier(sup_id: str, sup_in: SupplierUpdate, db: Session = Depends(get_db)):
    sup = db.query(Supplier).filter(Supplier.id == sup_id).first()
    if not sup:
        raise HTTPException(status_code=404, detail="Nhà cung cấp không tồn tại")
    
    sup.name = sup_in.name
    sup.contact_name = sup_in.contact_name
    sup.phone = sup_in.phone
    sup.email = sup_in.email
    sup.address = sup_in.address
    sup.is_active = sup_in.is_active
    
    db.commit()
    db.refresh(sup)
    return sup

@router.delete("/suppliers/{sup_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_supplier(sup_id: str, db: Session = Depends(get_db)):
    sup = db.query(Supplier).filter(Supplier.id == sup_id).first()
    if not sup:
        raise HTTPException(status_code=404, detail="Nhà cung cấp không tồn tại")
    db.delete(sup)
    db.commit()
    return None

@router.post("/suppliers/seed", status_code=status.HTTP_201_CREATED)
def seed_suppliers(db: Session = Depends(get_db)):
    """Tạo dữ liệu nhà cung cấp mẫu để test"""
    existing = db.query(Supplier).count()
    if existing > 0:
        return {"message": f"Đã có {existing} nhà cung cấp, không cần seed thêm."}
    
    sample_suppliers = [
        Supplier(id=str(uuid.uuid4()), name="Công ty TNHH Logitech Việt Nam", contact_name="Nguyễn Văn An", phone="0901234567", email="contact@logitech.vn", address="123 Nguyễn Huệ, Q.1, TP.HCM", is_active=True),
        Supplier(id=str(uuid.uuid4()), name="Razer Distributor Asia", contact_name="Trần Thị Bình", phone="0987654321", email="sales@razer.asia", address="456 Lê Lợi, Q.1, TP.HCM", is_active=True),
        Supplier(id=str(uuid.uuid4()), name="Corsair Vietnam", contact_name="Lê Minh Châu", phone="0912345678", email="info@corsair.vn", address="789 Điện Biên Phủ, Q.Bình Thạnh, TP.HCM", is_active=True),
        Supplier(id=str(uuid.uuid4()), name="SteelSeries Official Store", contact_name="Phạm Đức Duy", phone="0923456789", email="partner@steelseries.com", address="321 Cách Mạng Tháng 8, Q.3, TP.HCM", is_active=True),
        Supplier(id=str(uuid.uuid4()), name="HyperX / Kingston Technology", contact_name="Võ Thị Em", phone="0934567890", email="hyperx@kingston.vn", address="654 Võ Văn Tần, Q.3, TP.HCM", is_active=True),
    ]
    
    for sup in sample_suppliers:
        db.add(sup)
    db.commit()
    return {"message": f"Đã tạo {len(sample_suppliers)} nhà cung cấp mẫu thành công!"}


# =======================
# INVENTORY OVERVIEW
# =======================

@router.get("/skus", response_model=List[InventorySKUResponse])
def get_inventory_skus(db: Session = Depends(get_db)):
    skus = db.query(ProductSKU).options(joinedload(ProductSKU.product).selectinload(Product.images)).all()
    res = []
    for sku in skus:
        # Get first image of product if any
        img_url = None
        if sku.product.images:
            img_url = sku.product.images[0].url
            
        res.append(InventorySKUResponse(
            sku_id=sku.id,
            product_name=sku.product.name,
            sku_code=sku.sku_code,
            stock_quantity=sku.stock_quantity,
            price=sku.price,
            image_url=img_url,
            brand_id=sku.product.brand_id
        ))
    return res


@router.get("/skus/{sku_id}/history", response_model=List[SKUHistoryResponse])
def get_sku_history(sku_id: str, db: Session = Depends(get_db)):
    items = db.query(StockReceiptItem).join(StockReceipt).filter(StockReceiptItem.sku_id == sku_id).order_by(StockReceipt.created_at.desc()).all()
    res = []
    for item in items:
        res.append(SKUHistoryResponse(
            receipt_code=item.receipt.receipt_code,
            type=item.receipt.type,
            quantity=item.quantity,
            unit_price=item.unit_price,
            total_price=item.total_price,
            created_at=item.receipt.created_at,
            created_by=item.receipt.created_by,
            note=item.receipt.note
        ))
    return res

# =======================
# STOCK RECEIPTS (IN/OUT)
# =======================

@router.get("/receipts", response_model=List[StockReceiptResponse])
def get_all_receipts(db: Session = Depends(get_db)):
    return db.query(StockReceipt).order_by(StockReceipt.created_at.desc()).all()

@router.post("/receipts", response_model=StockReceiptResponse, status_code=status.HTTP_201_CREATED)
def create_receipt(receipt_in: StockReceiptCreate, current_admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    # Generate receipt code
    date_str = datetime.now().strftime("%Y%m%d")
    count = db.query(StockReceipt).filter(StockReceipt.created_at >= datetime.now().replace(hour=0, minute=0, second=0)).count()
    code = f"{receipt_in.type}-{date_str}-{(count + 1):03d}"

    total_amount = sum([item.quantity * item.unit_price for item in receipt_in.items])

    new_receipt = StockReceipt(
        id=str(uuid.uuid4()),
        receipt_code=code,
        type=receipt_in.type,
        supplier_id=receipt_in.supplier_id if receipt_in.type == 'IN' else None,
        total_amount=total_amount,
        note=receipt_in.note,
        created_by=current_admin.full_name
    )
    db.add(new_receipt)
    db.flush()

    for item_in in receipt_in.items:
        # Update SKU stock
        sku = db.query(ProductSKU).filter(ProductSKU.id == item_in.sku_id).first()
        if not sku:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"SKU {item_in.sku_id} không tồn tại")
        
        if receipt_in.type == 'IN':
            sku.stock_quantity += item_in.quantity
        elif receipt_in.type == 'OUT':
            if sku.stock_quantity < item_in.quantity:
                db.rollback()
                raise HTTPException(status_code=400, detail=f"Số lượng tồn kho không đủ để xuất cho SKU {item_in.sku_id}")
            sku.stock_quantity -= item_in.quantity

        receipt_item = StockReceiptItem(
            id=str(uuid.uuid4()),
            receipt_id=new_receipt.id,
            sku_id=item_in.sku_id,
            quantity=item_in.quantity,
            unit_price=item_in.unit_price,
            total_price=item_in.quantity * item_in.unit_price
        )
        db.add(receipt_item)

    db.commit()
    db.refresh(new_receipt)
    return new_receipt


@router.get("/receipts/{receipt_id}/export-excel")
def export_receipt_excel(receipt_id: str, db: Session = Depends(get_db)):
    """Xuất phiếu nhập/xuất kho dưới dạng file Excel theo mẫu TT133"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    receipt = db.query(StockReceipt).filter(StockReceipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="Phiếu không tồn tại")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phiếu Kho"

    # Styles
    bold_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    # Header info
    ws['A1'] = 'Đơn vị: Cửa hàng EZ4ENCE Gaming Gear'
    ws['A1'].font = bold_font
    ws['A2'] = 'Địa chỉ: TP.HCM'
    ws['A2'].font = normal_font

    # Title
    receipt_title = "PHIẾU NHẬP KHO" if receipt.type == 'IN' else "PHIẾU XUẤT KHO"
    ws.merge_cells('A4:G4')
    ws['A4'] = receipt_title
    ws['A4'].font = title_font
    ws['A4'].alignment = center

    ws.merge_cells('A5:G5')
    ws['A5'] = '(Ban hành theo Thông tư số 133/2016/TT-BTC ngày 26/8/2016 của Bộ Tài chính)'
    ws['A5'].font = Font(italic=True, size=9)
    ws['A5'].alignment = center

    # Receipt info
    created = receipt.created_at.strftime("%d/%m/%Y %H:%M") if receipt.created_at else ""
    ws.merge_cells('A6:G6')
    ws['A6'] = f'Ngày {receipt.created_at.strftime("%d")} tháng {receipt.created_at.strftime("%m")} năm {receipt.created_at.strftime("%Y")}'
    ws['A6'].alignment = center
    ws['A6'].font = Font(italic=True, size=10)

    ws.merge_cells('A7:G7')
    ws['A7'] = f'Số: {receipt.receipt_code}'
    ws['A7'].alignment = center
    ws['A7'].font = Font(bold=True, size=11)

    row = 9
    if receipt.type == 'IN' and receipt.supplier:
        ws[f'A{row}'] = f'Nhà cung cấp: {receipt.supplier.name}'
        ws[f'A{row}'].font = normal_font
        row += 1
    ws[f'A{row}'] = f'Người tạo phiếu: {receipt.created_by}'
    ws[f'A{row}'].font = normal_font
    row += 1
    if receipt.note:
        ws[f'A{row}'] = f'Ghi chú: {receipt.note}'
        ws[f'A{row}'].font = normal_font
        row += 1
    
    row += 1

    # Table header
    headers = ['STT', 'Tên hàng hóa, sản phẩm', 'Mã SKU', 'ĐVT', 'Số lượng', 'Đơn giá (VNĐ)', 'Thành tiền (VNĐ)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # Table body
    for idx, item in enumerate(receipt.items, 1):
        row += 1
        sku = db.query(ProductSKU).filter(ProductSKU.id == item.sku_id).first()
        product_name = sku.product.name if sku else "N/A"
        sku_code = sku.sku_code if sku else "N/A"

        data = [idx, product_name, sku_code, 'Cái', item.quantity, item.unit_price, item.total_price]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in [1, 3, 4, 5]:
                cell.alignment = center
            elif col_idx in [6, 7]:
                cell.number_format = '#,##0'

    # Total row
    row += 1
    ws.cell(row=row, column=1, value='').border = thin_border
    total_cell = ws.cell(row=row, column=2, value='TỔNG CỘNG')
    total_cell.font = Font(bold=True, size=10)
    total_cell.border = thin_border
    for col_idx in [3, 4]:
        ws.cell(row=row, column=col_idx, value='').border = thin_border
    qty_total = sum(item.quantity for item in receipt.items)
    ws.cell(row=row, column=5, value=qty_total).border = thin_border
    ws.cell(row=row, column=5).font = Font(bold=True, size=10)
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=6, value='').border = thin_border
    ws.cell(row=row, column=7, value=receipt.total_amount).border = thin_border
    ws.cell(row=row, column=7).font = Font(bold=True, size=10)
    ws.cell(row=row, column=7).number_format = '#,##0'

    # Signature section
    row += 3
    sigs = ['Người lập phiếu', 'Người giao hàng', 'Thủ kho', 'Kế toán trưởng', 'Giám đốc'] if receipt.type == 'IN' else ['Người lập phiếu', 'Người nhận hàng', 'Thủ kho', 'Giám đốc']
    
    # Spread signatures across columns
    if len(sigs) == 5:
        cols = [1, 2, 3, 5, 7]
    else:
        cols = [1, 3, 5, 7]
    
    for i, sig in enumerate(sigs):
        col = cols[i] if i < len(cols) else cols[-1]
        cell = ws.cell(row=row, column=col, value=sig)
        cell.font = Font(bold=True, size=10)
        cell.alignment = center
        cell2 = ws.cell(row=row+1, column=col, value='(Ký, họ tên)')
        cell2.font = Font(italic=True, size=9)
        cell2.alignment = center

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{receipt.receipt_code}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

